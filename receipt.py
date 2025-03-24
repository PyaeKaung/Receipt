import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter import PhotoImage
from tkcalendar import Calendar
import openpyxl
from openpyxl.styles import Font
import os
from PIL import ImageGrab
from PIL import Image, ImageTk
import win32com.client
import win32gui
from datetime import datetime
import time

def on_select(option):
    selected_format.set(option)

def clear_cells(ws):
    cells_to_clear = ["D3", "O3", "G14", "C17", "F17", "I17", "F19", "I19", "F21", "I21", "H23", "E26", "E27"]
    for cell in cells_to_clear:
        ws[cell].value = None

from datetime import datetime

def generate_files():
    format_selected = selected_format.get()
    template_path = "Template.xlsx"
    output_path = os.path.join(os.getcwd(), "Output")

    if not os.path.exists(output_path):
        os.makedirs(output_path)

    serial_number_input = serial_entry.get()
    received_from_input = received_from_entry.get()
    selected_date_str = calendar.get_date()
    selected_bank_input = selected_bank.get()
    intinput = input_entry.get()
    words_output = output_entry.get(1.0, tk.END).strip()
    transaction_details = transaction_details_entry.get("1.0", tk.END).strip()

    charge_selected = selected_left_option.get()
    transaction_selected = selected_right_option.get()
    combined_charges_transaction = f"{charge_selected} {transaction_selected}".strip()

    selected_date = datetime.strptime(selected_date_str, "%m/%d/%y")
    formatted_serial_number = serial_number_input.zfill(1)

    currency = selected_currency.get()

    if format_selected == "PNG":
        sstem_path = os.path.join(output_path, "sstem.xlsx")
        if os.path.exists(sstem_path):
            wb = openpyxl.load_workbook(sstem_path)
            ws = wb.active
            
            clear_cells(ws)

            if selected_option.get() == "BAGAN THANDE":
                ws["D3"].value = "ü"
                ws["D3"].font = Font(name='Wingdings', size=22)
            elif selected_option.get() == "THANDE BEACH":
                ws["O3"].value = "ü"
                ws["O3"].font = Font(name='Wingdings', size=22)

            ws["S10"].value = f'=IFERROR("{formatted_serial_number}", "")'
            ws["R11"].value = selected_date
            ws["R11"].number_format = "DD/MM/YYYY"
            ws["G14"].value = received_from_input
            ws["I17"].value = selected_bank_input
            
            ws["H23"].value = combined_charges_transaction

            try:
                numeric_value = float(intinput)
            except ValueError:
                numeric_value = None
            
            if currency == "Kyats":
                ws["F21"].value = numeric_value
                ws["I21"].value = words_output
            elif currency == "Dollars":
                ws["F19"].value = numeric_value
                ws["I19"].value = words_output

            ws["D24"].value = transaction_details
            
            wb.save(sstem_path)

            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            wb = excel.Workbooks.Open(sstem_path)
            excel.WindowState = -4137

            excel_hwnd = win32gui.FindWindow("XLMAIN", None)
            if excel_hwnd:
                win32gui.SetForegroundWindow(excel_hwnd)

            time.sleep(1)
            left, top, right, bottom = win32gui.GetWindowRect(excel_hwnd)

            a1_x = left + 50
            a1_y = top + 177
            s29_x = left + 970
            s29_y = top + 847

            screenshot_path = os.path.join(output_path, f"Receipt({formatted_serial_number}).jpg")
            screenshot = ImageGrab.grab(bbox=(a1_x, a1_y, s29_x, s29_y))
            screenshot.save(screenshot_path)
            
            wb.Close(SaveChanges=False)
            excel.Quit()

            messagebox.showinfo("Success", f"Screenshot saved: {screenshot_path}")

        else:
            messagebox.showerror("File Not Found", "sstem.xlsx not found in the Output folder.")

    elif format_selected == "Excel":
        output_excel_path = os.path.join(output_path, f"Receipt({formatted_serial_number}).xlsx")
        wb = openpyxl.load_workbook(template_path)

        ws = wb.active
        clear_cells(ws)

        if selected_option.get() == "BAGAN THANDE":
            ws["D3"].value = "ü"
            ws["D3"].font = Font(name='Wingdings', size=22)
        elif selected_option.get() == "THANDE BEACH":
            ws["O3"].value = "ü"
            ws["O3"].font = Font(name='Wingdings', size=22)

        ws["S10"].value = f'=IFERROR("{formatted_serial_number}", "")'
        ws["R11"].value = selected_date
        ws["R11"].number_format = "DD/MM/YYYY"
        ws["G14"].value = received_from_input
        ws["I17"].value = selected_bank_input
        
        ws["H23"].value = combined_charges_transaction

        try:
            numeric_value = float(intinput)
        except ValueError:
            numeric_value = None
        
        if currency == "Kyats":
            ws["F21"].value = numeric_value
            ws["I21"].value = words_output
        elif currency == "Dollars":
            ws["F19"].value = numeric_value
            ws["I19"].value = words_output

        ws["D24"].value = transaction_details

        try:
            wb.save(output_excel_path)
            messagebox.showinfo("Success", f"Excel file copied and modified: {output_excel_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving Excel file: {e}")

    else:
        messagebox.showwarning("Select Format", "Please select a format.")

    next_serial_number = str(int(serial_number_input) + 1).zfill(len(serial_number_input))
    serial_entry.delete(0, tk.END)
    serial_entry.insert(0, next_serial_number)

root = tk.Tk()
root.title("Receipt")
root.resizable(False, False)
root.geometry("1100x600")
root.configure(bg="white")

background_image_path = "background.jpg"
background_image = Image.open(background_image_path)
background_image = background_image.resize((1100, 600))
background_photo = ImageTk.PhotoImage(background_image)

background_label = tk.Label(root, image=background_photo)
background_label.place(x=0, y=0, relwidth=1, relheight=1)

selected_format = tk.StringVar()
selected_format.set("Select Format")

dropdown_button = tk.Button(
    root,
    textvariable=selected_format,
    bg="black",
    fg="white",
    font=("Helvetica", 12, "bold"),
    width=15,
    height=2,
    command=lambda: dropdown_menu.post(root.winfo_x() + 30, root.winfo_y() + 100)
)

dropdown_button.place(x=20, y=20)

dropdown_menu = tk.Menu(root, tearoff=0, bg="black", fg="white", font=("Helvetica", 12, "bold"))
dropdown_menu.add_command(label="PNG", command=lambda: on_select("PNG"))
dropdown_menu.add_command(label="Excel", command=lambda: on_select("Excel"))

radio_frame = tk.Frame(root, bg="white")
radio_frame.place(x=200, y=20)

selected_option = tk.StringVar(value="BAGAN THANDE")

radio_button1 = tk.Radiobutton(radio_frame, text="BAGAN THANDE", variable=selected_option, value="BAGAN THANDE", bg="white", font=("Helvetica", 12))
radio_button1.pack(anchor='w')

radio_button2 = tk.Radiobutton(radio_frame, text="THANDE BEACH", variable=selected_option, value="THANDE BEACH", bg="white", font=("Helvetica", 12))
radio_button2.pack(anchor='w')

serial_label = tk.Label(root, text="SERIAL NUMBER:", bg="white", fg="black", font=("Helvetica", 16, "bold"))
serial_label.place(x=800, y=20)

serial_entry = tk.Entry(root, bg="black", fg="white", font=("Helvetica", 20))
serial_entry.place(x=800, y=50, width=200)

received_from_label = tk.Label(root, text="RECEIVED FROM:", bg="white", fg="black", font=("Helvetica", 16, "bold"))
received_from_label.place(x=20, y=80)

received_from_entry = tk.Entry(root, bg="black", fg="white", font=("Helvetica", 20))
received_from_entry.place(x=20, y=110, width=200)

selected_bank = tk.StringVar()
selected_bank.set("Select Bank")

bank_dropdown_button = tk.Button(
    root,
    textvariable=selected_bank,
    bg="black",
    fg="white",
    font=("Helvetica", 12, "bold"),
    width=15,
    height=2,
    command=lambda: bank_dropdown_menu.post(root.winfo_x() + 30, root.winfo_y() + 240)
)

bank_dropdown_button.place(x=20, y=160)

bank_dropdown_menu = tk.Menu(root, tearoff=0, bg="black", fg="white", font=("Helvetica", 12, "bold"))
bank_dropdown_menu.add_command(label="KPAY", command=lambda: on_select_bank("KPAY"))
bank_dropdown_menu.add_command(label="AYA(Co.,)", command=lambda: on_select_bank("AYA(Co.,)"))
bank_dropdown_menu.add_command(label="KBZ(Co.,)", command=lambda: on_select_bank("KBZ(Co.,)"))
bank_dropdown_menu.add_command(label="CB(Co.,)", command=lambda: on_select_bank("CB(Co.,)"))

def on_select_bank(option):
    selected_bank.set(option)

input_label = tk.Label(root, text="IN NUMERAL:", bg="white", fg="black", font=("Helvetica", 14, "bold"))
input_label.place(x=20, y=230)

input_entry = tk.Entry(root, bg="black", fg="white", font=("Helvetica", 20))
input_entry.place(x=20, y=260, width=250)

output_label = tk.Label(root, text="IN WORD:", bg="white", fg="black", font=("Helvetica", 14, "bold"))
output_label.place(x=300, y=160)

output_entry = tk.Text(root, bg="lightgray", fg="black", font=("Helvetica", 20), height=3, wrap='word')
output_entry.place(x=300, y=190, width=450)

def number_to_words(n):
    """Convert a number into words and append 'Kyats' with proper capitalization."""
    if n < 0:
        return "negative " + number_to_words(-n).capitalize() + " Kyats"
    elif n == 0:
        return "zero Kyats"

    units = ["", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine"]
    teens = ["ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen", "seventeen", "eighteen", "nineteen"]
    tens = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]
    thousands = ["", "thousand", "million"]

    words = []
    if n >= 1000000:
        words.append(number_to_words(n // 1000000) + " million")  # Handle millions
        n %= 1000000
    if n >= 1000:
        words.append(number_to_words(n // 1000) + " thousand")  # Handle thousands
        n %= 1000
    if n >= 100:
        words.append(units[n // 100] + " hundred")
        n %= 100
    if n >= 20:
        words.append(tens[n // 10])
        n %= 10
    elif n >= 10:
        words.append(teens[n - 10])
        n = 0
    if n > 0:
        words.append(units[n])


    result = ' '.join(filter(bool, words)).strip().capitalize()
    return result


selected_currency = tk.StringVar()
selected_currency.set("Kyats")

currency_dropdown_button = tk.Button(
    root,
    textvariable=selected_currency,
    bg="black",
    fg="white",
    font=("Helvetica", 12, "bold"),
    width=15,
    height=2,
    command=lambda: currency_dropdown_menu.post(root.winfo_x() + 30, root.winfo_y() + 380)
)

currency_dropdown_button.place(x=20, y=300)

currency_dropdown_menu = tk.Menu(root, tearoff=0, bg="black", fg="white", font=("Helvetica", 12, "bold"))
currency_dropdown_menu.add_command(label="Kyats", command=lambda: on_select_currency("Kyats"))
currency_dropdown_menu.add_command(label="Dollars", command=lambda: on_select_currency("Dollars"))

left_dropdown_label = tk.Label(root, text="Charges:", bg="white", fg="black", font=("Helvetica", 14, "bold"))
left_dropdown_label.place(x=20, y=370)

selected_left_option = tk.StringVar()
selected_left_option.set("")

left_dropdown_button = tk.Button(
    root,
    textvariable=selected_left_option,
    bg="black",
    fg="white",
    font=("Helvetica", 12, "bold"),
    width=15,
    height=2,
    command=lambda: left_dropdown_menu.post(root.winfo_x() + 30, root.winfo_y() + 480)
)

left_dropdown_button.place(x=20, y=400)

left_dropdown_menu = tk.Menu(root, tearoff=0, bg="black", fg="white", font=("Helvetica", 12, "bold"))
left_dropdown_menu.add_command(label="Room Charges", command=lambda: selected_left_option.set("Room Charges"))
left_dropdown_menu.add_command(label="Other Charges", command=lambda: selected_left_option.set("Other Charges"))

right_dropdown_label = tk.Label(root, text="Transaction:", bg="white", fg="black", font=("Helvetica", 14, "bold"))
right_dropdown_label.place(x=230, y=370)

selected_right_option = tk.StringVar()
selected_right_option.set("")

right_dropdown_button = tk.Button(
    root,
    textvariable=selected_right_option,
    bg="black",
    fg="white",
    font=("Helvetica", 12, "bold"),
    width=15,
    height=2,
    command=lambda: right_dropdown_menu.post(root.winfo_x() + 230, root.winfo_y() + 460)
)

right_dropdown_button.place(x=230, y=400)

right_dropdown_menu = tk.Menu(root, tearoff=0, bg="black", fg="white", font=("Helvetica", 12, "bold"))
right_dropdown_menu.add_command(label="-", command=lambda: selected_right_option.set(" "))
right_dropdown_menu.add_command(label="(Deposit)", command=lambda: selected_right_option.set("(Deposit)"))
right_dropdown_menu.add_command(label="(Balance)", command=lambda: selected_right_option.set("(Balance)"))

transaction_details_label = tk.Label(root, text="Remarks:", bg="white", fg="black", font=("Helvetica", 14, "bold"))
transaction_details_label.place(x=470, y=320)

transaction_details_entry = tk.Text(root, bg="lightgray", fg="black", font=("Helvetica", 12), height=6, wrap='word')
transaction_details_entry.place(x=470, y=350, width=450)

def on_select_currency(option):
    selected_currency.set(option)

def update_output():
    try:
        number = int(input_entry.get())
        currency = selected_currency.get()
        words = number_to_words(number) + f" {currency}"
        capitalized_words = words.title()
        output_entry.config(state='normal')
        output_entry.delete(1.0, tk.END)
        output_entry.insert(tk.END, capitalized_words)
        output_entry.config(state='disabled')
    except ValueError:
        output_entry.config(state='normal')
        output_entry.delete(1.0, tk.END)
        output_entry.config(state='disabled')

input_entry.bind("<KeyRelease>", lambda event: update_output())

received_from_label.place(x=20, y=80)
received_from_entry.place(x=20, y=110, width=200)

received_from_label.place(x=20, y=80)
received_from_entry.place(x=20, y=110, width=200)

calendar_label = tk.Label(root, text="Select Date:", bg="white", fg="black", font=("Helvetica", 14, "bold"))
calendar_label.place(x=800, y=90)

calendar = Calendar(root, selectmode='day', year=datetime.now().year, month=datetime.now().month, day=datetime.now().day) 
calendar.place(x=800, y=120)

button = tk.Button(
    root,
    text="Generate",
    bg="blue",
    fg="white",
    font=("Helvetica", 12, "bold"),
    width=15,
    height=3,
    command=generate_files
)
button.place(x=230, y=500)

root.mainloop()